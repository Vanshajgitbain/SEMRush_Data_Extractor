#!/usr/bin/env python
"""
SEMRUSH Data Extractor - Streamlit App
Extracts chart data from Semrush and other analytics platforms
"""

import streamlit as st
import time
import re
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Set page config
st.set_page_config(
    page_title="SEMRUSH Data Extractor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .chart-card {
        border: 2px solid #e0e0e0;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
        background-color: #f9f9f9;
        cursor: pointer;
        transition: all 0.3s;
    }
    .chart-card:hover {
        border-color: #dc143c;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .progress-section {
        background-color: #f0f0f0;
        padding: 20px;
        border-radius: 8px;
        margin: 20px 0;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'driver' not in st.session_state:
    st.session_state.driver = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = 'intro'  # intro, browser_opened, detecting, charts, extracting, results
if 'charts' not in st.session_state:
    st.session_state.charts = []
if 'selected_chart' not in st.session_state:
    st.session_state.selected_chart = None
if 'tooltips' not in st.session_state:
    st.session_state.tooltips = []
if 'extraction_complete' not in st.session_state:
    st.session_state.extraction_complete = False
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None


def create_driver():
    """Create a Chrome driver with anti-detection settings."""
    try:
        options = Options()
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        return driver
    except Exception as e:
        raise Exception(f"Failed to create WebDriver: {str(e)}")


def is_driver_valid(driver):
    """Check if driver session is still valid."""
    if driver is None:
        return False
    try:
        # Try a simple command to check if session is valid
        driver.current_window_handle
        return True
    except:
        return False


def cleanup_driver():
    """Safely close the WebDriver."""
    if st.session_state.driver is not None:
        try:
            st.session_state.driver.quit()
        except:
            pass
        st.session_state.driver = None


def find_charts(driver):
    """Find chart sections on the page by looking for headings near SVGs."""
    charts = driver.execute_script("""
    var results = [];
    var svgs = document.querySelectorAll('svg');
    var seen = new Set();
    
    for (var svg of svgs) {
        var r = svg.getBoundingClientRect();
        if (r.width < 200 || r.height < 80) continue;
        
        var parent = svg.parentElement;
        var title = '';
        for (var i = 0; i < 10 && parent; i++) {
            var headings = parent.querySelectorAll('h2, h3, h4, [class*="title"], [class*="Title"]');
            for (var h of headings) {
                var t = h.textContent.trim();
                if (t.length > 2 && t.length < 80 && !seen.has(t)) {
                    title = t;
                    break;
                }
            }
            if (title) break;
            parent = parent.parentElement;
        }
        
        if (!title) title = 'Chart (' + Math.round(r.width) + 'x' + Math.round(r.height) + ')';
        
        var key = title.replace(/(.{4,})\\1/i, '$1').trim();
        if (seen.has(key)) continue;
        seen.add(key);
        
        var absY = window.scrollY + r.top;
        results.push({
            title: key,
            width: Math.round(r.width),
            height: Math.round(r.height),
            y: Math.round(absY)
        });
    }
    return results;
    """)
    return charts


def find_chart_svg(driver, chart_info):
    """Find the SVG element for a selected chart."""
    y_pos = chart_info['y']
    
    svg = driver.execute_script("""
    var targetY = arguments[0];
    var allSvgs = document.querySelectorAll('svg');
    var best = null;
    var bestDist = Infinity;
    
    for (var s of allSvgs) {
        var r = s.getBoundingClientRect();
        if (r.width < 200 || r.height < 80) continue;
        var absY = window.scrollY + r.top;
        var dist = Math.abs(absY - targetY);
        if (dist < 200 && dist < bestDist) {
            bestDist = dist;
            best = s;
        }
    }
    
    if (best) best.scrollIntoView({block: 'center'});
    return best;
    """, y_pos)
    
    return svg


def extract_tooltips(driver, svg, progress_bar=None):
    """Hover across the chart SVG to capture tooltip text."""
    tooltips = []
    is_semrush_chart = False
    
    svg_width = svg.size['width']
    svg_height = svg.size['height']
    
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", svg)
    time.sleep(1)
    
    # Quick activation
    try:
        driver.execute_script("""
        var svg = arguments[0];
        var rect = svg.getBoundingClientRect();
        var x = rect.left + rect.width / 2;
        var y = rect.top + rect.height / 2;
        var target = document.elementFromPoint(x, y) || svg;
        ['pointerenter', 'mouseover'].forEach(function(evtName) {
            target.dispatchEvent(new PointerEvent(evtName, {
                clientX: x, clientY: y,
                bubbles: true, cancelable: true, view: window
            }));
        });
        """, svg)
        time.sleep(0.5)
    except:
        pass
    
    y_offsets = [0, -int(svg_height * 0.05), -int(svg_height * 0.1), -int(svg_height * 0.15), -int(svg_height * 0.2), 
                 int(svg_height * 0.05), int(svg_height * 0.1)]
    
    # SMART PROBE
    probe_positions = list(range(0, 101, 10))
    data_found_positions = []
    
    for probe_i in probe_positions:
        x_frac = probe_i / 100
        x_off = int(-svg_width + svg_width * 2 * x_frac)
        
        found_at_this_x = False
        for y_off in y_offsets:
            driver.execute_script("""
            var svg = arguments[0];
            var xOff = arguments[1];
            var yOff = arguments[2];
            var rect = svg.getBoundingClientRect();
            var x = rect.left + rect.width/2 + xOff;
            var y = rect.top + rect.height/2 + yOff;
            var target = document.elementFromPoint(x, y) || svg;
            ['pointerenter','pointermove','mouseover','mousemove'].forEach(function(evtName) {
                target.dispatchEvent(new PointerEvent(evtName, {
                    clientX: x, clientY: y,
                    bubbles: true, cancelable: true, view: window
                }));
            });
            """, svg, x_off, y_off)
            time.sleep(0.08)
            
            found = driver.execute_script("""
            var results = [];
            var selectors = ['[role="tooltip"]', 'div[class*="tooltip"]', 'div[class*="Tooltip"]',
                            'div[class*="popover"]', 'div[class*="Popover"]', 'div[class*="chartTooltip"]'];
            for (var sel of selectors) {
                try {
                    var els = document.querySelectorAll(sel);
                    for (var el of els) {
                        var style = window.getComputedStyle(el);
                        if (style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0') {
                            var text = el.textContent.trim();
                            if (text.length > 10 && text.length < 800) return true;
                        }
                    }
                } catch(e) {}
            }
            return false;
            """)
            
            if found:
                data_found_positions.append(probe_i)
                found_at_this_x = True
                break
    
    if not data_found_positions:
        return tooltips
    
    min_pos = max(0, min(data_found_positions) - 8)
    max_pos = min(100, max(data_found_positions) + 8)
    
    num_positions = (max_pos - min_pos) + 1
    
    for i in range(min_pos, max_pos + 1):
        try:
            x_frac = i / 100
            x_off = int(-svg_width + svg_width * 2 * x_frac)
            
            for y_off in y_offsets:
                driver.execute_script("""
                var svg = arguments[0];
                var xOff = arguments[1];
                var yOff = arguments[2];
                var rect = svg.getBoundingClientRect();
                var x = rect.left + rect.width/2 + xOff;
                var y = rect.top + rect.height/2 + yOff;
                var target = document.elementFromPoint(x, y) || svg;
                ['pointerenter','pointermove','mouseover','mousemove'].forEach(function(evtName) {
                    target.dispatchEvent(new PointerEvent(evtName, {
                        clientX: x, clientY: y,
                        bubbles: true, cancelable: true, view: window
                    }));
                });
                """, svg, x_off, y_off)
                time.sleep(0.15)
                
                found = driver.execute_script("""
                var results = [];
                var selectors = [
                    '[role="tooltip"]',
                    'div[class*="tooltip"]', 'div[class*="Tooltip"]',
                    'div[class*="popover"]', 'div[class*="Popover"]',
                    'div[class*="chartTooltip"]', 'div[class*="chart-tooltip"]',
                    'g[role="tooltip"]', 'text[class*="tooltip"]'
                ];
                var checked = new Set();
                for (var sel of selectors) {
                    try {
                        var els = document.querySelectorAll(sel);
                        for (var el of els) {
                            if (checked.has(el)) continue;
                            checked.add(el);
                            var style = window.getComputedStyle(el);
                            if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') continue;
                            var text = el.textContent.trim();
                            if (text.length > 10 && text.length < 800) {
                                results.push(text);
                            }
                        }
                    } catch(e) {}
                }
                if (results.length === 0) {
                    var divs = document.querySelectorAll('div');
                    for (var d of divs) {
                        if (checked.has(d)) continue;
                        var r = d.getBoundingClientRect();
                        if (r.width > 50 && r.width < 500 && r.height > 30 && r.height < 500) {
                            var st = window.getComputedStyle(d);
                            if (st.position === 'absolute' || st.position === 'fixed') {
                                if (st.display === 'none' || st.visibility === 'hidden' || st.opacity === '0') continue;
                                var t = d.textContent.trim();
                                if (t.length > 10 && t.length < 800 && /\\d/.test(t)) {
                                    results.push(t);
                                }
                            }
                        }
                    }
                }
                return results;
                """)
                
                if found:
                    got_one = False
                    for tip in found:
                        has_domain = bool(re.search(r'[a-z0-9\-]+\.com', tip, re.IGNORECASE))
                        has_metrics = bool(re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Mon|Tue|Wed|Thu|Fri|Sat|Sun)', tip, re.IGNORECASE))
                        is_valid = False
                        
                        if has_domain and 'Difference' not in tip:
                            is_valid = True
                            is_semrush_chart = True
                        elif has_metrics and 'Difference' not in tip:
                            is_valid = True
                        
                        if is_valid and tip not in tooltips:
                            tooltips.append(tip)
                            got_one = True
                    if got_one:
                        break
                    
            if progress_bar:
                progress = (i - min_pos) / num_positions
                progress_bar.progress(min(progress, 0.95))
                
        except Exception as e:
            pass
    
    return tooltips


def parse_metrics_tooltips(tooltips):
    """Parse metrics-style tooltips for charts like Traffic Trend."""
    rows = []
    MONTHS = r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec'
    monthly_re = re.compile(r'(' + MONTHS + r')\s+(\d{4})')
    
    for tip in tooltips:
        # Try to find period (date) in the tooltip
        monthly_match = monthly_re.search(tip)
        
        if not monthly_match:
            continue
            
        period = f"{monthly_match.group(1)} {monthly_match.group(2)}"
        
        # Remove the period from the tooltip to get the rest
        remaining = tip[monthly_match.end():].strip()
        
        # Extract domain/metric name (text before the first digit)
        metric_match = re.match(r'^([a-zA-Z\.\-]+)(\d+[kmKM]?)', remaining)
        
        if metric_match:
            metric_name = metric_match.group(1).strip()
            # Extract the numeric value with unit
            value_match = re.search(r'(\d+\.?\d*[kmKM]?)', remaining)
            
            if value_match:
                metric_value = value_match.group(1).strip()
                
                if metric_name and metric_value:
                    rows.append({
                        'period': period,
                        'metric': metric_name,
                        'value': metric_value
                    })
    
    if not rows:
        return None
    
    # Deduplicate
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r['period'], r['metric'], r['value'])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    
    periods_set = set(r['period'] for r in unique_rows)
    metrics_set = set(r['metric'] for r in unique_rows)
    
    month_order = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                   'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    
    def period_sort_key(p):
        year_m = re.search(r'(\d{4})', p)
        year = int(year_m.group(1)) if year_m else 0
        mon_m = re.match(r'(\w+)', p)
        mon = month_order.get(mon_m.group(1), 0) if mon_m else 0
        return (year, mon)
    
    sorted_periods = sorted(periods_set, key=period_sort_key)
    sorted_metrics = sorted(metrics_set)
    
    return {
        'rows': unique_rows,
        'periods': sorted_periods,
        'metrics': sorted_metrics,
        'type': 'metrics'
    }


def create_excel_file(tooltips, filename="chart_data"):
    """Create Excel file from tooltips."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Data"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Try to parse structured data
    data = parse_metrics_tooltips(tooltips)
    
    if data and data['type'] == 'metrics':
        periods = data['periods']
        metrics = data['metrics']
        rows = data['rows']
        
        if periods and metrics:
            pivot = {}
            for r in rows:
                p = r['period']
                m = r['metric']
                pivot.setdefault(p, {})[m] = r['value']
            
            headers = ['Period'] + metrics
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="dc143c", end_color="dc143c", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            for row_idx, p in enumerate(periods, 2):
                cell = ws.cell(row=row_idx, column=1, value=p)
                cell.border = thin_border
                cell.font = Font(bold=True)
                
                for col_idx, m in enumerate(metrics, 2):
                    val = pivot.get(p, {}).get(m, '-')
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
            
            for col_idx, h in enumerate(headers, 1):
                max_len = len(str(h))
                for row_idx in range(2, len(periods) + 2):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3
    else:
        # Fallback: Create a raw data sheet if parsing fails
        # Clear existing content by deleting all rows
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        ws.append(['Extracted Data'])
        ws.append([''])  # Empty row
        ws.append(['The following data was captured from the tooltips:'])
        ws.append([''])  # Empty row
        
        for idx, tooltip in enumerate(tooltips, 1):
            ws.append([f'Data Point {idx}', tooltip])
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# Main UI
def main():
    # Header with Bain Logo and Title
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        try:
            st.image("bain_logo.png", width=100)
        except:
            st.markdown("**BAIN & COMPANY**", unsafe_allow_html=True)
    with col2:
        st.markdown("<h1 style='text-align: center; color: #dc143c;'>📊 SEMRUSH Data Extractor</h1>", unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Two-column layout: Left (About) + Right (Steps)
    left_col, right_col = st.columns([1, 2.5])
    
    # LEFT COLUMN: About This Platform (Always Visible)
    with left_col:
        with st.expander("ℹ️ About This Platform", expanded=True):
            st.markdown("""
            **SEMRUSH Data Extractor** is an intelligent automation tool that extracts chart data from Semrush analytics pages and similar data visualization platforms. 
            Instead of manually copying data point by point, this tool automatically finds all charts on a page, hovers over them to trigger tooltips, intelligently captures all data points, 
            and exports the results to professionally formatted Excel files.
            
            ---
            
            **Step-by-Step Process:**
            
            1. **Enter URL** - Provide URL
            2. **Run & Scan** - Open page
            3. **Confirm Ready** - Verify ready
            4. **View Charts** - See charts
            5. **Select Chart** - Pick chart
            6. **Monitor Progress** - Track progress
            7. **Download Data** - Get Excel
            """)
    
    # RIGHT COLUMN: All Workflow Steps
    with right_col:
        # STEP 1: Enter Data Source URL (Always visible)
        st.markdown("<h3 style='color: #dc143c;'>Step 1: Enter Data Source URL</h3>", unsafe_allow_html=True)
        url = st.text_input(
            "Enter the URL you want to extract data from:",
            value="https://www.semrush.com/analytics/adwords/positions/",
            key="url_input"
        )
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("🚀 Run", use_container_width=True, key="run_button"):
                if url:
                    st.session_state.url = url
                    st.session_state.current_page = 'browser_opened'
                    st.rerun()
                else:
                    st.error("Please enter a URL")
        
        # STEP 2: Browser Opened
        if st.session_state.current_page in ['browser_opened', 'ready_to_detect', 'detecting', 'charts', 'extracting', 'results']:
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 2: Browser Opened</h3>", unsafe_allow_html=True)
            
            status_container = st.container()
            
            if st.session_state.current_page == 'browser_opened':
                # LOADING STATE
                status_container.info("🌐 Opening URL in browser...")
                
                try:
                    with status_container:
                        st.write("Creating WebDriver...")
                    
                    if st.session_state.driver is None or not is_driver_valid(st.session_state.driver):
                        cleanup_driver()
                        st.session_state.driver = create_driver()
                    
                    with status_container:
                        st.write("Opening URL in browser...")
                    
                    driver = st.session_state.driver
                    url = st.session_state.url
                    
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    
                    try:
                        driver.get(url)
                    except:
                        time.sleep(1)
                        driver.get(url)
                    
                    with status_container:
                        st.write("Waiting for page to load...")
                    
                    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    time.sleep(5)
                    
                    with status_container:
                        st.success("✅ Page loaded successfully!")
                    
                    time.sleep(1)
                    st.session_state.current_page = 'ready_to_detect'
                    st.rerun()
                        
                except Exception as e:
                    status_container.error(f"Error opening URL: {str(e)}")
                    time.sleep(2)
                    cleanup_driver()
                    st.session_state.current_page = 'intro'
                    st.rerun()
            else:
                # COMPLETED STATE
                st.success("✅ Browser opened and page loaded")
        
        # STEP 3: Confirm Ready to Detect
        if st.session_state.current_page in ['ready_to_detect', 'detecting', 'charts', 'extracting', 'results']:
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 3: Confirm Ready for Detection</h3>", unsafe_allow_html=True)
            
            if st.session_state.current_page == 'ready_to_detect':
                st.write("📍 The browser has opened and the page is fully loaded.")
                st.write("Click **Yes** to start detecting charts on this page.")
                
                col1, col2 = st.columns([1, 1])
                with col1:
                    if st.button("✅ Yes, Detect Charts", use_container_width=True, key="detect_yes"):
                        st.session_state.current_page = 'detecting'
                        st.rerun()
                with col2:
                    if st.button("🔙 Back", use_container_width=True, key="back_btn"):
                        cleanup_driver()
                        st.session_state.current_page = 'intro'
                        st.rerun()
            else:
                st.success("✅ Ready to detect - proceeding with detection")
        
        # STEP 4: Detecting Charts
        if st.session_state.current_page in ['detecting', 'charts', 'extracting', 'results']:
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 4: Detecting Charts</h3>", unsafe_allow_html=True)
            
            if st.session_state.current_page == 'detecting':
                st.info("🔍 Scanning for charts...")
                status_container = st.container()
                
                try:
                    if not is_driver_valid(st.session_state.driver):
                        st.error("Browser session was lost. Please start over.")
                        time.sleep(2)
                        cleanup_driver()
                        st.session_state.current_page = 'intro'
                        st.rerun()
                    
                    driver = st.session_state.driver
                    
                    with status_container:
                        st.write("Scanning page for charts...")
                    
                    charts = find_charts(driver)
                    st.session_state.charts = charts
                    
                    if charts:
                        with status_container:
                            st.success(f"✅ Found {len(charts)} charts!")
                        time.sleep(1)
                        st.session_state.current_page = 'charts'
                        st.rerun()
                    else:
                        with status_container:
                            st.warning("No charts found on this page.")
                        time.sleep(2)
                        st.session_state.current_page = 'ready_to_detect'
                        st.rerun()
                        
                except Exception as e:
                    status_container.error(f"Error during chart detection: {str(e)}")
                    time.sleep(2)
                    cleanup_driver()
                    st.session_state.current_page = 'intro'
                    st.rerun()
            else:
                st.success(f"✅ Found {len(st.session_state.charts)} charts")
        
        # STEP 5: Select a Chart
        if st.session_state.current_page in ['charts', 'extracting', 'results']:
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 5: Select a Chart to Extract</h3>", unsafe_allow_html=True)
            
            if st.session_state.current_page == 'charts':
                st.write(f"Found {len(st.session_state.charts)} charts - select one below:")
                
                cols = st.columns(2)
                for idx, chart in enumerate(st.session_state.charts):
                    with cols[idx % 2]:
                        st.markdown(f"""
                        <div style='border: 1px solid #e0e0e0; padding: 10px; border-radius: 5px;'>
                            <h5>{chart['title']}</h5>
                            <p style='font-size: 12px; color: gray;'>Dimensions: {chart['width']}x{chart['height']}px</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if st.button(f"📊 Extract", key=f"extract_{idx}", use_container_width=True):
                            st.session_state.selected_chart = chart
                            st.session_state.current_page = 'extracting'
                            st.rerun()
            elif st.session_state.selected_chart is not None:
                st.success(f"✅ Selected: {st.session_state.selected_chart['title']}")
        
        # STEP 6: Extracting Data
        if st.session_state.current_page in ['extracting', 'results']:
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 6: Extracting Data</h3>", unsafe_allow_html=True)
            
            if st.session_state.current_page == 'extracting':
                chart = st.session_state.selected_chart
                st.info(f"📊 Extracting from: {chart['title']}")
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    if not is_driver_valid(st.session_state.driver):
                        st.error("Browser session was lost. Please start over.")
                        time.sleep(2)
                        cleanup_driver()
                        st.session_state.current_page = 'intro'
                        st.rerun()
                    
                    driver = st.session_state.driver
                    
                    status_text.write("Locating chart...")
                    svg = find_chart_svg(driver, chart)
                    
                    if not svg:
                        st.error("Could not locate the chart SVG element.")
                        time.sleep(2)
                        st.session_state.current_page = 'charts'
                        st.rerun()
                    
                    status_text.write("Extracting data points...")
                    tooltips = extract_tooltips(driver, svg, progress_bar)
                    
                    st.session_state.tooltips = tooltips
                    st.session_state.extraction_complete = True
                    
                    progress_bar.progress(1.0)
                    status_text.success(f"✅ Extraction complete! Found {len(tooltips)} data points.")
                    
                    time.sleep(1)
                    st.session_state.current_page = 'results'
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"Extraction error: {str(e)}")
                    time.sleep(2)
                    st.session_state.current_page = 'charts'
                    st.rerun()
            else:
                st.success(f"✅ Extracted {len(st.session_state.tooltips)} data points")
        
        # STEP 7: Download Data
        if st.session_state.current_page == 'results':
            st.markdown("---")
            st.markdown("<h3 style='color: #dc143c;'>Step 7: Download Your Data</h3>", unsafe_allow_html=True)
            
            chart = st.session_state.selected_chart
            tooltips = st.session_state.tooltips
            
            if chart and tooltips:
                st.success(f"✅ Extracted {len(tooltips)} data points from '{chart['title']}'")
            else:
                st.error("Error: Missing chart or tooltip data")
                st.stop()
            
            col1, col2 = st.columns([3, 1])
            with col1:
                filename = st.text_input(
                    "Enter the name for your Excel file:",
                    value="chart_data",
                    key="filename_input",
                    placeholder="chart_data"
                )
            
            with col2:
                st.write("")
                if filename:
                    try:
                        excel_file = create_excel_file(tooltips, filename)
                        st.download_button(
                            label="📥 Download",
                            data=excel_file,
                            file_name=f"{filename}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_button",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
            
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("📊 Extract Another Chart", use_container_width=True, key="another_chart"):
                    st.session_state.current_page = 'charts'
                    st.rerun()
            
            with col2:
                if st.button("🏠 Start Over", use_container_width=True, key="start_over"):
                    cleanup_driver()
                    st.session_state.current_page = 'intro'
                    st.session_state.charts = []
                    st.session_state.tooltips = []
                    st.session_state.selected_chart = None
                    st.rerun()


if __name__ == "__main__":
    main()
