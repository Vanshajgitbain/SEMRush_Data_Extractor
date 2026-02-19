#!/usr/bin/env python
"""
Simple Chart Tooltip Extractor
Opens a URL, finds charts, hovers to capture tooltips, prints a table.
"""

import time
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


def create_driver():
    """Create a Chrome driver with anti-detection settings."""
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def find_charts(driver):
    """Find chart sections on the page by looking for headings near SVGs."""
    charts = driver.execute_script("""
    var results = [];
    var svgs = document.querySelectorAll('svg');
    var seen = new Set();
    
    for (var svg of svgs) {
        var r = svg.getBoundingClientRect();
        if (r.width < 200 || r.height < 80) continue;
        
        // Walk up to find a heading/title for this chart
        var parent = svg.parentElement;
        var title = '';
        for (var i = 0; i < 10 && parent; i++) {
            var headings = parent.querySelectorAll('h2, h3, h4, [class*="title"], [class*="Title"]');
            for (var h of headings) {
                var t = h.textContent.trim();
                if (t.length > 2 && t.length < 80 && !seen.has(t)) {
                    title = t;
                    break;
                }
            }
            if (title) break;
            parent = parent.parentElement;
        }
        
        if (!title) title = 'Chart (' + Math.round(r.width) + 'x' + Math.round(r.height) + ')';
        
        // Deduplicate
        var key = title.replace(/(.{4,})\\1/i, '$1').trim();
        if (seen.has(key)) continue;
        seen.add(key);
        
        var absY = window.scrollY + r.top;
        results.push({
            title: key,
            width: Math.round(r.width),
            height: Math.round(r.height),
            y: Math.round(absY)
        });
    }
    return results;
    """)
    return charts


def find_chart_svg(driver, chart_info):
    """Find the SVG element for a selected chart."""
    y_pos = chart_info['y']
    
    svg = driver.execute_script("""
    var targetY = arguments[0];
    var allSvgs = document.querySelectorAll('svg');
    var best = null;
    var bestDist = Infinity;
    
    for (var s of allSvgs) {
        var r = s.getBoundingClientRect();
        if (r.width < 200 || r.height < 80) continue;
        var absY = window.scrollY + r.top;
        var dist = Math.abs(absY - targetY);
        if (dist < 200 && dist < bestDist) {
            bestDist = dist;
            best = s;
        }
    }
    
    if (best) best.scrollIntoView({block: 'center'});
    return best;
    """, y_pos)
    
    return svg


def extract_tooltips(driver, svg):
    """Hover across the chart SVG to capture tooltip text."""
    tooltips = []
    is_semrush_chart = False  # Flag to detect chart type
    
    svg_width = svg.size['width']
    svg_height = svg.size['height']
    
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", svg)
    time.sleep(1)
    
    # Quick activation: dispatch a single event to activate tooltip system without traversing
    try:
        driver.execute_script("""
        var svg = arguments[0];
        var rect = svg.getBoundingClientRect();
        var x = rect.left + rect.width / 2;
        var y = rect.top + rect.height / 2;
        var target = document.elementFromPoint(x, y) || svg;
        ['pointerenter', 'mouseover'].forEach(function(evtName) {
            target.dispatchEvent(new PointerEvent(evtName, {
                clientX: x, clientY: y,
                bubbles: true, cancelable: true, view: window
            }));
        });
        """, svg)
        time.sleep(0.5)
    except:
        pass
    
    # Try multiple Y offsets to find where tooltips trigger
    y_offsets = [0, -int(svg_height * 0.05), -int(svg_height * 0.1), -int(svg_height * 0.15), -int(svg_height * 0.2), 
                 int(svg_height * 0.05), int(svg_height * 0.1)]
    
    # SMART PROBE: Find the active data region with all Y offsets
    print(f"[*] Probing chart to find active data region...")
    probe_positions = list(range(0, 101, 10))  # Every 10% (0, 10, 20, ..., 100)
    data_found_positions = []
    
    for probe_i in probe_positions:
        x_frac = probe_i / 100
        x_off = int(-svg_width + svg_width * 2 * x_frac)
        
        found_at_this_x = False
        # Try all Y offsets for each X position
        for y_off in y_offsets:
            driver.execute_script("""
            var svg = arguments[0];
            var xOff = arguments[1];
            var yOff = arguments[2];
            var rect = svg.getBoundingClientRect();
            var x = rect.left + rect.width/2 + xOff;
            var y = rect.top + rect.height/2 + yOff;
            var target = document.elementFromPoint(x, y) || svg;
            ['pointerenter','pointermove','mouseover','mousemove'].forEach(function(evtName) {
                target.dispatchEvent(new PointerEvent(evtName, {
                    clientX: x, clientY: y,
                    bubbles: true, cancelable: true, view: window
                }));
            });
            """, svg, x_off, y_off)
            time.sleep(0.08)
            
            found = driver.execute_script("""
            var results = [];
            var selectors = ['[role="tooltip"]', 'div[class*="tooltip"]', 'div[class*="Tooltip"]',
                            'div[class*="popover"]', 'div[class*="Popover"]', 'div[class*="chartTooltip"]'];
            for (var sel of selectors) {
                try {
                    var els = document.querySelectorAll(sel);
                    for (var el of els) {
                        var style = window.getComputedStyle(el);
                        if (style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0') {
                            var text = el.textContent.trim();
                            if (text.length > 10 && text.length < 800) return true;
                        }
                    }
                } catch(e) {}
            }
            return false;
            """)
            
            if found:
                data_found_positions.append(probe_i)
                found_at_this_x = True
                break  # Found data at this X, move to next X position
    
    # Determine scan range based on findings
    if not data_found_positions:
        print("[!] No data points detected in any area.")
        print(f"[+] Extraction completed: 0 tooltips captured")
        return tooltips
    
    min_pos = max(0, min(data_found_positions) - 8)  # Start 8 before first data found
    max_pos = min(100, max(data_found_positions) + 8)  # End 8 after last data found
    
    print(f"[*] Active data region detected: positions {min_pos}-{max_pos}")
    
    # Detailed sweep only in the active region
    num_positions = (max_pos - min_pos) + 1
    print(f"[*] Sweeping {num_positions} positions in active area ({svg_width}x{svg_height}px)...")
    
    for i in range(min_pos, max_pos + 1):
        try:
            x_frac = i / 100
            # Use range: -100% to +100% for efficient coverage with 50 positions
            x_off = int(-svg_width + svg_width * 2 * x_frac)
            
            # Try each Y offset until we get a tooltip
            for y_off in y_offsets:
                # Dispatch pointer/mouse events via JS for React compatibility (primary method)
                driver.execute_script("""
                var svg = arguments[0];
                var xOff = arguments[1];
                var yOff = arguments[2];
                var rect = svg.getBoundingClientRect();
                var x = rect.left + rect.width/2 + xOff;
                var y = rect.top + rect.height/2 + yOff;
                var target = document.elementFromPoint(x, y) || svg;
                ['pointerenter','pointermove','mouseover','mousemove'].forEach(function(evtName) {
                    target.dispatchEvent(new PointerEvent(evtName, {
                        clientX: x, clientY: y,
                        bubbles: true, cancelable: true, view: window
                    }));
                });
                """, svg, x_off, y_off)
                time.sleep(0.15)
                
                found = driver.execute_script("""
                var results = [];
                var selectors = [
                    '[role="tooltip"]',
                    'div[class*="tooltip"]', 'div[class*="Tooltip"]',
                    'div[class*="popover"]', 'div[class*="Popover"]',
                    'div[class*="chartTooltip"]', 'div[class*="chart-tooltip"]',
                    'g[role="tooltip"]', 'text[class*="tooltip"]'  // SVG tooltips
                ];
                var checked = new Set();
                for (var sel of selectors) {
                    try {
                        var els = document.querySelectorAll(sel);
                        for (var el of els) {
                            if (checked.has(el)) continue;
                            checked.add(el);
                            var style = window.getComputedStyle(el);
                            if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') continue;
                            var text = el.textContent.trim();
                            if (text.length > 10 && text.length < 800) {
                                results.push(text);
                            }
                        }
                    } catch(e) {}
                }
                // Fallback: look for absolutely-positioned divs with data
                if (results.length === 0) {
                    var divs = document.querySelectorAll('div');
                    for (var d of divs) {
                        if (checked.has(d)) continue;
                        var r = d.getBoundingClientRect();
                        if (r.width > 50 && r.width < 500 && r.height > 30 && r.height < 500) {
                            var st = window.getComputedStyle(d);
                            if (st.position === 'absolute' || st.position === 'fixed') {
                                if (st.display === 'none' || st.visibility === 'hidden' || st.opacity === '0') continue;
                                var t = d.textContent.trim();
                                if (t.length > 10 && t.length < 800 && /\\d/.test(t)) {
                                    results.push(t);
                                }
                            }
                        }
                    }
                }
                return results;
                """)
                
                if found:
                    got_one = False
                    for tip in found:
                        # Detect chart type: check if tooltip has .com domains (Semrush) or metrics (stats chart)
                        has_domain = bool(re.search(r'[a-z0-9\-]+\.com', tip, re.IGNORECASE))
                        has_metrics = bool(re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Mon|Tue|Wed|Thu|Fri|Sat|Sun)', tip, re.IGNORECASE))
                        is_valid = False
                        
                        # Semrush charts: must have .com and no "Difference"
                        if has_domain and 'Difference' not in tip:
                            is_valid = True
                            is_semrush_chart = True
                        # Metrics/stats charts: have dates or common metric names
                        elif has_metrics and 'Difference' not in tip:
                            is_valid = True
                        
                        if is_valid and tip not in tooltips:
                            tooltips.append(tip)
                            got_one = True
                    if got_one:
                        break  # Found tooltip at this Y offset, move to next X
                    
            if (i - min_pos) % max(1, (max_pos - min_pos) // 4) == 0:
                print(f"   Position {i}/{max_pos} -- {len(tooltips)} unique tooltips so far")
                
        except Exception as e:
            pass
    
    print(f"[+] Extraction completed: {len(tooltips)} total tooltips captured")
    return tooltips


def parse_and_print_table(tooltips):
    """Parse tooltip text and print a formatted table."""
    if not tooltips:
        print("\n[!] No tooltips captured.")
        return
    
    print(f"\n{'='*70}")
    print(f"RAW TOOLTIPS CAPTURED ({len(tooltips)})")
    print(f"{'='*70}")
    for i, tip in enumerate(tooltips, 1):
        preview = tip[:200].replace('\n', ' | ')
        print(f"  {i}. {preview}")
    
    # Detect chart type: Semrush (has .com domains) or Metrics (has dates with values)
    is_semrush = any(re.search(r'[a-z0-9\-]+\.com', tip, re.IGNORECASE) for tip in tooltips)
    
    if is_semrush:
        # SEMRUSH-STYLE PARSING (keep existing logic)
        parse_semrush_tooltips(tooltips)
    else:
        # METRICS-STYLE PARSING (for charts like Traffic Trend)
        parse_metrics_tooltips(tooltips)


def parse_semrush_tooltips(tooltips):
    """Parse Semrush-style tooltips with domain data."""
    # Parse Semrush-style tooltips:
    # Monthly format: "Nov 2025 hm.com 13.5M (12.5M - 15.6M) zara.com ..."
    # Daily format:   "Sat, Jan 17, 2026hm.com156.9K(146.7K – 173.6K)..."
    # Weekly format:  "Dec 29, 2025 – Jan 4, 2026hm.com1.2M..." or "Jan 12 – 18hm.com..."
    rows = []
    MONTHS = r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec'
    # Match weekly range: "Dec 29, 2025 – Jan 4, 2026" or "Jan 12 – 18" or "Jan 26 – Feb 1"
    weekly_re = re.compile(
        r'(' + MONTHS + r')\s+(\d{1,2}),?\s*(\d{4})?\s*'
        r'[–\-]\s*'
        r'(?:(' + MONTHS + r')\s+)?(\d{1,2}),?\s*(\d{4})?'
    )
    # Match daily: "Mon, Jan 17, 2026"
    daily_re = re.compile(
        r'(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),?\s*'
        r'(' + MONTHS + r')\s+(\d{1,2}),?\s*(\d{4})'
    )
    monthly_re = re.compile(r'(' + MONTHS + r')\s+(\d{4})')
    
    for tip in tooltips:
        # Strip noise text from forecasts
        tip_clean = re.sub(r'Forecast\s+based\s+on\s+previous\s+available\s+data\.?\s*Updated\s+weekly\.?', '', tip)
        
        # Extract the period (try weekly first, then daily, then monthly)
        weekly_match = weekly_re.search(tip_clean)
        daily_match = daily_re.search(tip_clean)
        monthly_match = monthly_re.search(tip_clean)
        
        if weekly_match:
            start_mon = weekly_match.group(1)
            start_day = weekly_match.group(2)
            start_year = weekly_match.group(3)
            end_mon = weekly_match.group(4)  # may be None
            end_day = weekly_match.group(5)
            end_year = weekly_match.group(6)  # may be None
            # Resolve year: use end_year or start_year or infer from context
            year = end_year or start_year or '2026'
            # If start month is Dec and end month is Jan, start year = end year - 1
            if start_year is None and end_year:
                start_year = end_year
                if start_mon == 'Dec' and end_mon and end_mon == 'Jan':
                    start_year = str(int(end_year) - 1)
            elif start_year is None:
                start_year = year
            if end_mon:
                period = f"{start_mon} {start_day} – {end_mon} {end_day}"
            else:
                period = f"{start_mon} {start_day} – {end_day}"
            # Store start year for sorting (attach hidden)
            period = f"{period}, {start_year}"
        elif daily_match:
            period = f"{daily_match.group(1)} {daily_match.group(2)}, {daily_match.group(3)}"
        elif monthly_match:
            period = f"{monthly_match.group(1)} {monthly_match.group(2)}"
        else:
            continue
        
        # Find all company.com entries and the bold value right after them
        # Pattern: domain.com followed by a value like 13.5M or 156.9K
        company_pattern = re.compile(
            r'([a-z0-9\-]+\.com)'
            r'[\s:]*'
            r'(\d{1,3}(?:[,.]\d+)?\s*[MKBmkb])'
            r'[\s]*'
            r'(?:\([^)]*\))?',
            re.IGNORECASE
        )
        
        matches = company_pattern.findall(tip_clean)
        for domain, value in matches:
            domain = domain.lower().strip()
            # Clean domain: strip leading digits (e.g. "2025hm.com" -> "hm.com")
            # and "forecast" prefix (e.g. "forecasthm.com" -> "hm.com")
            domain = re.sub(r'^\d+', '', domain)
            domain = re.sub(r'^forecast', '', domain, flags=re.IGNORECASE)
            value = value.strip()
            if domain and domain not in ('google.com', 'semrush.com'):
                rows.append({'period': period, 'entity': domain, 'value': value})
    
    if not rows:
        print("\n[!] Could not parse structured data from tooltips.")
        print("    The raw tooltip text is shown above.")
        return
    
    # Deduplicate
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r['entity'], r['period'], r['value'])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    
    month_order = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                   'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    
    # Build pivot: {period -> {company -> value}}
    periods_set = set()
    companies_set = set()
    pivot = {}
    for r in unique_rows:
        p = r['period']
        c = r['entity']
        periods_set.add(p)
        companies_set.add(c)
        pivot.setdefault(p, {})[c] = r['value']
    
    # Sort periods chronologically
    def period_sort_key(p):
        # Weekly: "Jan 12 – 18, 2026" or "Dec 29 – Jan 4, 2025"
        # Daily: "Jan 17, 2026"
        # Monthly: "Nov 2025"
        # Extract first month, first day, and year from the period string
        # Try to find year (last 4-digit number)
        year_m = re.search(r'(\d{4})', p)
        year = int(year_m.group(1)) if year_m else 0
        # Extract first month name
        mon_m = re.match(r'(\w+)', p)
        mon = month_order.get(mon_m.group(1), 0) if mon_m else 0
        # Extract first day number
        day_m = re.search(r'[A-Za-z]+\s+(\d{1,2})', p)
        day = int(day_m.group(1)) if day_m else 0
        return (year, mon, day)
    
    sorted_periods = sorted(periods_set, key=period_sort_key)
    
    # Sort companies: put the main/largest one first, rest alphabetically
    # Heuristic: sort alphabetically but group by typical naming
    sorted_companies = sorted(companies_set)
    
    # Build summary line
    num_periods = len(sorted_periods)
    # Detect period type for label
    has_range = any('–' in p or '-' in p for p in sorted_periods)
    has_comma = any(',' in p for p in sorted_periods)
    period_label = "weeks" if has_range else "dates" if has_comma else "months"
    period_names = [sorted_periods[-2], sorted_periods[-1]] if num_periods >= 2 else sorted_periods[-1:]
    inc_text = " and ".join(period_names) if len(period_names) == 2 else period_names[0] if period_names else ""
    
    print(f"\nExtracted {num_periods} {period_label} including {inc_text}:\n")
    
    # Column widths
    pw = max(12, max(len(p) for p in sorted_periods) + 2)
    col_widths = {}
    for c in sorted_companies:
        max_val_len = max((len(pivot.get(p, {}).get(c, '-')) for p in sorted_periods), default=3)
        col_widths[c] = max(len(c) + 2, max_val_len + 2)
    
    # Print header
    header = f"{'Period':<{pw}}"
    for c in sorted_companies:
        header += f"{c:<{col_widths[c]}}"
    
    sep = '-' * len(header)
    print(header)
    print(sep)
    
    # Print rows
    for p in sorted_periods:
        row = f"{p:<{pw}}"
        for c in sorted_companies:
            val = pivot.get(p, {}).get(c, '-')
            row += f"{val:<{col_widths[c]}}"
        print(row)
    
    print(sep)
    print(f"Total: {len(unique_rows)} data points")
    
    # --- Export to Excel ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chart Data"
        
        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Write header row
        headers = ["Period"] + sorted_companies
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Write data rows
        for row_idx, p in enumerate(sorted_periods, 2):
            cell = ws.cell(row=row_idx, column=1, value=p)
            cell.border = thin_border
            for col_idx, c in enumerate(sorted_companies, 2):
                val = pivot.get(p, {}).get(c, '-')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        
        # Auto-fit column widths
        for col_idx, h in enumerate(headers, 1):
            max_len = len(str(h))
            for row_idx in range(2, len(sorted_periods) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3
        
        # Save with unique filename (timestamp)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(os.getcwd(), f"chart_data_{timestamp}.xlsx")
        wb.save(excel_path)
        print(f"\n[+] Excel saved: {excel_path}")
    except Exception as e:
        print(f"\n[!] Could not save Excel: {e}")


def parse_metrics_tooltips(tooltips):
    """Parse metrics/stats chart tooltips (like Traffic Trend with dates and values)."""
    rows = []
    MONTHS = r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec'
    
    # Match multiple date formats:
    # 1. Daily: "Mon, Jan 17, 2026" or "Jan 17, 2026"
    # 2. Short month+day: "Mar 25" (no year)
    # 3. Monthly: "Mar 2024" (4-digit year)
    daily_re = re.compile(
        r'(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),?\s*'
        r'(' + MONTHS + r')\s+(\d{1,2}),?\s*(\d{4})?'
    )
    short_date_re = re.compile(
        r'(' + MONTHS + r')\s+(\d{1,2})(?!\d)'  # Month + 1-2 digit day (not followed by more digits)
    )
    monthly_re = re.compile(
        r'(' + MONTHS + r')\s+(\d{4})'
    )
    
    for tip in tooltips:
        # Skip forecast data - only parse actual data
        if 'Forecast' in tip:
            continue
        
        # Try to extract date (daily first, then short date, then monthly)
        daily_match = daily_re.search(tip)
        short_match = short_date_re.search(tip) if not daily_match else None
        monthly_match = monthly_re.search(tip) if not daily_match and not short_match else None
        
        period = None
        tip_clean = tip
        if daily_match:
            month = daily_match.group(1)
            day = daily_match.group(2)
            year = daily_match.group(3) or '2026'
            period = f"{month} {day}, {year}"
            tip_clean = daily_re.sub('', tip).strip()
        elif short_match:
            month = short_match.group(1)
            day = short_match.group(2)
            period = f"{month} {day}"
            tip_clean = short_date_re.sub('', tip).strip()
        elif monthly_match:
            month = monthly_match.group(1)
            year = monthly_match.group(2)
            period = f"{month} {year}"
            tip_clean = monthly_re.sub('', tip).strip()
        else:
            continue
        
        # Parse metrics: handle multiple formats
        # Format 1: Simple format "Traffic Cost $340,976.00" or "Traffic Cost: 340976"
        # Format 2: Complex format "Visits9.5535.5M" or "Visits 16.09% 398.6K"
        
        # First try simple format: MetricName[$value] or MetricName[: $value]
        simple_pattern = re.compile(
            r'([A-Za-z\s]+?)'  # Metric name (one or more words, non-greedy)
            r'[\s:]*'  # Optional colon/spaces
            r'([$]?[\d,]+\.?\d*)'  # Dollar value or plain number (with commas and decimals)
        )
        
        simple_matches = simple_pattern.findall(tip_clean)
        simple_found = False
        
        for metric_name, value in simple_matches:
            metric_name = metric_name.strip().lower().replace(' ', '_')
            value = value.strip()
            
            # Check if it looks like a valid value (has $ or is a reasonably large number)
            if metric_name and ('$' in value or (not value[0].isalpha())):
                rows.append({
                    'period': period, 
                    'metric': metric_name, 
                    'percentage': '',  # No percentage for simple format
                    'value': value
                })
                simple_found = True
        
        # If simple format didn't work, try complex format with percentages
        if not simple_found:
            # Match: metric name + optional spaces + number (percentage) + optional spaces/% 
            # + number with K/M/B unit (value)
            metric_pattern = re.compile(
                r'([A-Za-z\s]+?)'  # Metric name (one or more words, non-greedy)
                r'[\s]*'
                r'([\d.]+)'  # Percentage (just the number, % may not be present)
                r'[\s%]*'  # Optional % and spaces
                r'([\d.]+[KMB])'  # Absolute value with unit
                r'(?:\s*\([^)]*\))?'  # Optional range in parentheses
            )
            
            matches = metric_pattern.findall(tip_clean)
            for metric_name, percentage_num, value in matches:
                metric_name = metric_name.strip().lower().replace(' ', '_')
                # Ensure percentage is properly formatted (add % if missing)
                if '%' not in percentage_num:
                    percentage = f"{percentage_num}%"
                else:
                    percentage = percentage_num.strip()
                value = value.strip()
                
                if metric_name and percentage_num and value:
                    rows.append({
                        'period': period, 
                        'metric': metric_name, 
                        'percentage': percentage,
                        'value': value
                    })
    
    if not rows:
        print("\n[!] Could not parse metrics from tooltips.")
        print("    The raw tooltip text is shown above.")
        return
    
    # Deduplicate
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r['metric'], r['period'], r['percentage'], r['value'])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    
    month_order = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                   'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    
    # Build pivot: {period -> {metric -> {percentage: x, value: y}}}
    periods_set = set()
    metrics_set = set()
    pivot = {}
    for r in unique_rows:
        p = r['period']
        m = r['metric']
        periods_set.add(p)
        metrics_set.add(m)
        pivot.setdefault(p, {})[m] = {'percentage': r['percentage'], 'value': r['value']}
    
    # Sort periods chronologically
    def period_sort_key(p):
        year_m = re.search(r'(\d{4})', p)
        year = int(year_m.group(1)) if year_m else 0
        mon_m = re.search(r'(' + MONTHS + r')', p)
        mon = month_order.get(mon_m.group(1), 0) if mon_m else 0
        day_m = re.search(r'[A-Za-z]+\s+(\d{1,2})', p)
        day = int(day_m.group(1)) if day_m else 0
        return (year, mon, day)
    
    sorted_periods = sorted(periods_set, key=period_sort_key)
    sorted_metrics = sorted(metrics_set)
    
    print(f"\nExtracted {len(sorted_periods)} dates with {len(sorted_metrics)} metrics:\n")
    
    # Determine if we have percentages (complex format) or just values (simple format)
    has_percentages = any(r['percentage'] for r in unique_rows)
    
    # Print table
    pw = max(15, max(len(p) for p in sorted_periods) + 2)
    col_widths = {}
    
    if has_percentages:
        # Complex format: metrics with % and Value columns
        header = f"{'Date':<{pw}}"
        for m in sorted_metrics:
            col_widths[f"{m}_pct"] = max(len(m) + 4, 8)
            col_widths[f"{m}_val"] = max(len(m) + 4, 10)
            header += f"{m.upper()} %{' ' * (col_widths[f'{m}_pct'] - len(m) - 2)}{m.upper()} Value{' ' * (col_widths[f'{m}_val'] - len(m) - 6)}"
        
        sep = '-' * len(header)
        print(header)
        print(sep)
        
        # Data rows
        for p in sorted_periods:
            row = f"{p:<{pw}}"
            for m in sorted_metrics:
                data = pivot.get(p, {}).get(m, {})
                pct = data.get('percentage', '-')
                val = data.get('value', '-')
                row += f"{pct:<{col_widths[f'{m}_pct']}}{val:<{col_widths[f'{m}_val']}}"
            print(row)
    else:
        # Simple format: just metric and value columns
        header = f"{'Date':<{pw}}"
        for m in sorted_metrics:
            col_widths[m] = max(len(m) + 2, 12)
            header += f"{m.upper():<{col_widths[m]}}"
        
        sep = '-' * len(header)
        print(header)
        print(sep)
        
        # Data rows
        for p in sorted_periods:
            row = f"{p:<{pw}}"
            for m in sorted_metrics:
                data = pivot.get(p, {}).get(m, {})
                val = data.get('value', '-')
                row += f"{val:<{col_widths[m]}}"
            print(row)
    
    print(sep)
    print(f"Total: {len(unique_rows)} data points")
    
    # --- Export to Excel ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chart Data"
        
        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Build headers based on format
        if has_percentages:
            headers = ["Date"]
            for m in sorted_metrics:
                headers.append(f"{m.replace('_', ' ').title()} %")
                headers.append(f"{m.replace('_', ' ').title()} Value")
        else:
            headers = ["Date"] + [m.replace('_', ' ').title() for m in sorted_metrics]
        
        # Write header row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Write data rows
        for row_idx, p in enumerate(sorted_periods, 2):
            col_idx = 1
            cell = ws.cell(row=row_idx, column=col_idx, value=p)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
            col_idx += 1
            
            if has_percentages:
                for m in sorted_metrics:
                    data = pivot.get(p, {}).get(m, {})
                    pct = data.get('percentage', '-')
                    val = data.get('value', '-')
                    
                    # Percentage column
                    cell = ws.cell(row=row_idx, column=col_idx, value=pct)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    col_idx += 1
                    
                    # Value column (bold for absolute values)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    cell.font = Font(bold=True)
                    col_idx += 1
            else:
                for m in sorted_metrics:
                    data = pivot.get(p, {}).get(m, {})
                    val = data.get('value', '-')
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    cell.font = Font(bold=True)
                    col_idx += 1
        
        # Auto-fit column widths
        for col_idx, h in enumerate(headers, 1):
            max_len = len(str(h))
            for row_idx in range(2, len(sorted_periods) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3
        
        # Save with unique filename (timestamp)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(os.getcwd(), f"chart_data_{timestamp}.xlsx")
        wb.save(excel_path)
        print(f"\n[+] Excel saved: {excel_path}")
    except Exception as e:
        print(f"\n[!] Could not save Excel: {e}")


def main():
    print("\n" + "=" * 60)
    print("  CHART TOOLTIP EXTRACTOR")
    print("  Hover over charts -> capture tooltips -> print table")
    print("=" * 60 + "\n")
    
    url = input("[*] Paste URL:\n>>> ").strip()
    while not url:
        url = input(">>> ").strip()
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    driver = None
    try:
        print(f"\n[*] Opening: {url}")
        driver = create_driver()
        driver.get(url)
        print("[*] Waiting for page to load...")
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(5)
        
        input("\n[?] Are you ready to start the data extraction process? (Press Enter to continue): ")
        time.sleep(1)
        
        # Continuous extraction loop
        while True:
            print("\n[*] Scanning for charts...")
            charts = find_charts(driver)
            
            if not charts:
                print("[!] No charts found on this page.")
                break
            
            print(f"\n[+] Found {len(charts)} charts:\n")
            for i, c in enumerate(charts, 1):
                print(f"    {i}. {c['title']}  ({c['width']}x{c['height']}px)")
            
            while True:
                try:
                    choice = int(input(f"\n[?] Select chart (1-{len(charts)}): ").strip())
                    if 1 <= choice <= len(charts):
                        break
                except ValueError:
                    pass
                print(f"    Enter a number between 1 and {len(charts)}")
            
            selected = charts[choice - 1]
            print(f"\n[*] Selected: {selected['title']}")
            
            svg = find_chart_svg(driver, selected)
            if not svg:
                print("[!] Could not locate the chart SVG element.")
                break
            
            tooltips = extract_tooltips(driver, svg)
            parse_and_print_table(tooltips)
            
            # Ask if user wants more extractions
            print("\n" + "=" * 60)
            more = input("[?] Do you need more extractions? (y/n, default n): ").strip().lower()
            
            if more not in ('y', 'yes'):
                break
            
            # Ask if same page or new URL
            same_page = input("[?] Extract from same page or new URL? (same/new, default same): ").strip().lower()
            if same_page in ('new', 'n'):
                new_url = input("\n[*] Paste new URL:\n>>> ").strip()
                if new_url:
                    if not new_url.startswith(('http://', 'https://')):
                        new_url = 'https://' + new_url
                    print(f"\n[*] Opening: {new_url}")
                    driver.get(new_url)
                    print("[*] Waiting for page to load...")
                    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    time.sleep(3)
                    
                    input("\n[?] Are you ready to start the data extraction process? (Press Enter to continue): ")
                    time.sleep(1)
        
    except KeyboardInterrupt:
        print("\n[!] Extraction cancelled by user.")
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            input("\n[*] Press Enter to close browser...")
            driver.quit()


if __name__ == "__main__":
    main()
